#ライブラリ呼び出し
from selenium import webdriver #ウェブサイトを自動で操作するためのツール「selenium」から、「webdriver」を取り入れる
from selenium.webdriver.chrome.options import Options
from selenium.common.exceptions import NoSuchElementException
from webdriver_manager.chrome import ChromeDriverManager #ChromeDriverを自動でダウンロードしてくれるツール「webdriver_manager」から、「ChromeDriverManager」を取り入れる
from bs4 import BeautifulSoup #ウェブページの内容を解析するためのツール「BeautifulSoup」を取り入れています。
from time import sleep #「sleep」は、プログラムの実行を一時停止するためのツールです。
from datetime import datetime #「datetime」は、日付や時間を扱うためのツールです。
import re
import pandas as pd
import numpy as np 
import openpyxl #Excelファイルを操作するためのツール「openpyxl」を取り入れる
from openpyxl.styles import Font #「Font」は、Excelファイルのセルの書式を操作するためのツールです。
from openpyxl.styles import Border, Side


# ヘッドレスモードのオプションを設定する
chrome_options = Options()
chrome_options.add_argument("--headless")

# ChromeDriverManagerを使用してChromeドライバーを起動する
browser = webdriver.Chrome(ChromeDriverManager().install(), options=chrome_options)

# 出力するExcelファイル名
now = datetime.now()
timestamp_str = now.strftime("%Y%m%d%H%M")
filename = f'dlsite_sales_ranking_data_{timestamp_str}.xlsx'

# Excelファイルに書き込むためのライターを作成
with pd.ExcelWriter(filename, engine='openpyxl') as writer:
    # 巡回したいURLと対応するシート名
    url_sheet_mapping = {
        'https://www.dlsite.com/maniax/genre/list': 'ジャンル一覧(タグ)',
        'https://www.dlsite.com/maniax/ranking/day': '24時間(同人全体)',
        'https://www.dlsite.com/maniax/ranking/day?category=game&sub=RPG': '24時間(RPG)',
        'https://www.dlsite.com/maniax/ranking/day?category=game&sub=ADV': '24時間(ｱﾄﾞﾍﾞﾝﾁｬｰ)',
        'https://www.dlsite.com/maniax/ranking/day?category=game&sub=ACN': '24時間(ｱｸｼｮﾝ)',
        'https://www.dlsite.com/maniax/ranking/day?category=game&sub=SLN': '24時間(ｼﾐｭﾚｰｼｮﾝ)',
        'https://www.dlsite.com/maniax/ranking/week': '7日間(同人全体)',
        'https://www.dlsite.com/maniax/ranking/week?category=game&sub=RPG': '7日間(RPG)',
        'https://www.dlsite.com/maniax/ranking/week?category=game&sub=ADV': '7日間(ｱﾄﾞﾍﾞﾝﾁｬｰ)',
        'https://www.dlsite.com/maniax/ranking/week?category=game&sub=ACN': '7日間(ｱｸｼｮﾝ)',
        'https://www.dlsite.com/maniax/ranking/week?category=game&sub=SLN': '7日間(ｼﾐｭﾚｰｼｮﾝ)',
        'https://www.dlsite.com/maniax/ranking/month': '30日間(同人全体)',
        'https://www.dlsite.com/maniax/ranking/month?category=game&sub=RPG': '30日間(RPG)',
        'https://www.dlsite.com/maniax/ranking/month?category=game&sub=ADV': '30日間(ｱﾄﾞﾍﾞﾝﾁｬｰ)',
        'https://www.dlsite.com/maniax/ranking/month?category=game&sub=ACN': '30日間(ｱｸｼｮﾝ)',
        'https://www.dlsite.com/maniax/ranking/month?category=game&sub=SLN': '30日間(ｼﾐｭﾚｰｼｮﾝ)',
        'https://www.dlsite.com/maniax/ranking/year': '年間(同人全体)',
        'https://www.dlsite.com/maniax/ranking/year?category=game&sub=RPG': '年間(RPG)',
        'https://www.dlsite.com/maniax/ranking/year?category=game&sub=ADV': '年間(ｱﾄﾞﾍﾞﾝﾁｬｰ)',
        'https://www.dlsite.com/maniax/ranking/year?category=game&sub=ACN': '年間(ｱｸｼｮﾝ)',
        'https://www.dlsite.com/maniax/ranking/year?category=game&sub=SLN': '年間(ｼﾐｭﾚｰｼｮﾝ)',
        'https://www.dlsite.com/maniax/ranking/total': '累計(同人全体)',
        'https://www.dlsite.com/maniax/ranking/total?category=game&sub=RPG': '累計(RPG)',
        'https://www.dlsite.com/maniax/ranking/total?category=game&sub=ADV': '累計(ｱﾄﾞﾍﾞﾝﾁｬｰ)',
        'https://www.dlsite.com/maniax/ranking/total?category=game&sub=ACN': '累計(ｱｸｼｮﾝ)',
        'https://www.dlsite.com/maniax/ranking/total?category=game&sub=SLN': '累計(ｼﾐｭﾚｰｼｮﾝ)'

    }

    for url, sheet_name in url_sheet_mapping.items():   
        browser.get(url) # 指定したURLのページを開く
        sleep(1) # 1秒間待機する

        if sheet_name != 'ジャンル一覧(タグ)':
            try:
                elem_login_btn_top = browser.find_element_by_css_selector('.btn_yes.btn-approval') # ログインボタンの要素を指定する
                elem_login_btn_top.click() # ログインボタンをクリックする
            except NoSuchElementException:
                print(sheet_name +'のデータを作成します。R18以上チェックボタン確認不可。次に進みます')

            html_source = browser.page_source # ブラウザのページソースを取得する
            soup = BeautifulSoup(html_source, 'html.parser') # BeautifulSoupでブラウザのページソースを解析する
            
            #商品ページを1つに絞る
            data=[]
            products = soup.select('tr')
            for product in products:
                #ランキングナンバーの取得
                rank_elements = product.find('div', class_=lambda x: x and x.startswith('rank_no'))
                rank_no = int(rank_elements.text)  # 最初の要素のテキストを整数に変換

                #タイトルの取得
                title=product.select('a')[1]
                title=title.text.strip()

                #サークル名の取得
                circul_name=product.select('a')[2]
                circul_name=circul_name.text.strip()

                #製品ジャンルの取得
                category_name=product.find('div',attrs = {'class':'work_category'})
                category_name=category_name.text.strip()

                #販売数の取得
                dl_count_div = product.find('div', class_='dl_count')
                dl_count_text = dl_count_div.text
                dl_count = int(''.join(filter(str.isdigit, dl_count_text)))
                #filter 関数と str.isdigit メソッドを使って、文字列から数字だけを取り出します。
                #join メソッドを使って、取り出した数字を文字列に結合します。
                #int 関数を使って、文字列を整数型に変換します。

                #販売価格の取得
                work_price = product.find('span', class_=lambda x: x and x.startswith('work_price'))
                work_price_str = work_price.text.replace(',', '').replace('円', '') # テキスト内容を取得し、カンマを削除
                work_price_int = int(work_price_str) # 文字列を整数型に変換

                #登録売価の取得
                strike_price = product.find('span', attrs={'class': 'strike'})
                if strike_price is not None:
                    strike_price_str = strike_price.text.replace(',', '').replace('円', '')
                    strike_price_int = int(strike_price_str)
                else:
                    strike_price_int =  '※割引無'  
                
                #最低売上金額
                Sales = dl_count*work_price_int
                
                #最低粗利金額
                # 卸価格を計算する関数を定義します。この関数は販売価格に応じて卸価格を返します。(23年8月22日時点　https://www.dlsite.com/home/circle/regulations　に基づく)
                def calculate_strike_price(work_price_int):
                    # 以下の計算は同じままです
                    if work_price_int < 1000:
                        return work_price_int * 0.5
                    elif work_price_int < 2000:
                        return work_price_int * 0.6364
                    elif work_price_int < 3000:
                        return work_price_int * 0.7143
                    elif work_price_int < 4000:
                        return work_price_int * 0.7647
                    else:
                        return work_price_int * 0.8  # 販売価格4000円以上の場合、卸価格は販売価格の80%
                gross_Sales = int(dl_count * calculate_strike_price(work_price_int))

                #割引率の取得
                discount_percentage=product.find('span',attrs = {'class':'icon_campaign'}) # テキスト内容を取得し、%の前の数字部分だけを取り出す
                if discount_percentage is not None:
                    discount_percentage_str = discount_percentage.text.split('%')[0]
                    discount_percentage_int = int(discount_percentage_str)
                else:
                    discount_percentage_int= ''  # 空欄表記

                #割引終了時期の取得
                priod_date = product.find('span', attrs={'class': 'period_date'}) # テキスト内容を取得
                if priod_date is not None:
                    priod_date_text = priod_date.text
                    date_match = re.search(r'(\d{4})年(\d{2})月(\d{2})日 (\d{2})時(\d{2})分', priod_date_text)
                    #正規表現を使って、日付と時刻のパターンに一致する部分を検索します。この例では、正規表現の (\d{4})年(\d{2})月(\d{2})日 (\d{2})時(\d{2})分 が使用されています。
                    formatted_date = f'{date_match.group(1)}/{int(date_match.group(2))}/{date_match.group(3)}/{date_match.group(4)}:{date_match.group(5)}'
                    #re.search の結果から、各部分（年、月、日、時、分）を取り出し、所望のフォーマットに整形します。
                else:
                    formatted_date = ''  # 空欄表記

                #タグの取得
                search_tag = product.find('dd', class_='search_tag') # クラス名が"search_tag"であるddタグを探す
                if search_tag is not None:
                    tags = [a_tag.text for a_tag in search_tag.find_all('a')] # aタグのテキスト内容を取得し、一つの文字列に結合
                    output_string = '／'.join(tags) # タグを「／」で区切って一つの文字列に結合
                else:
                    output_string = ''  # 空欄表記

                #紹介文の取得
                work_text = product.find('dd', attrs={'class': 'work_text'}) 
                work_text=work_text.text

                #紹介文の文字数取得
                work_text_len=len(work_text)

                #「音声」「音楽」「体験版」などの有無
                work_genre_tag = product.find('dd', class_='work_genre') # クラス名が"work_genre"であるddタグを探す
                attributes = [span_tag.text for span_tag in work_genre_tag.find_all('span') if span_tag.text] # spanタグのテキスト内容を取得し、一つの文字列に結合
                work_genre_tag_string = '／'.join(attributes)

                #URL
                product_url_tag = product.find('a', class_='work_thumb_box')
                product_url = product_url_tag['href'] # href属性を取得

                details = {}
                datum=details
                datum['No.']=rank_no
                datum['商品名']=title
                datum['サークル名']=circul_name
                datum['ジャンル']=category_name
                datum['販売数']=dl_count 
                datum['現状価格']=work_price_int  
                datum['登録売価']=strike_price_int 
                datum['最低売上金額']= Sales
                datum['最低粗利金額']= gross_Sales
                datum['割引率']=discount_percentage_int
                datum['割引終了']=formatted_date
                datum['性癖関連のタグ']=output_string 
                datum['紹介文']=work_text
                datum['文字数']=work_text_len
                datum['音声等']=work_genre_tag_string 
                datum['URL']=product_url
                data.append(datum)

            df = pd.DataFrame(data)
            df.index = np.arange(1, len(df)+1)
            df.to_excel(writer, sheet_name,index=False)

        elif sheet_name == 'ジャンル一覧(タグ)':
            try:
                elem_login_btn_top = browser.find_element_by_css_selector('.btn_yes.btn-approval') # ログインボタンの要素を指定する
                elem_login_btn_top.click() # ログインボタンをクリックする
            except NoSuchElementException:
                print(sheet_name +'のデータを作成します。R18以上チェックボタン確認不可。次に進みます')

            html_source = browser.page_source # ブラウザのページソースを取得する
            soup = BeautifulSoup(html_source, 'html.parser') # BeautifulSoupでブラウザのページソースを解析する
           
            data = []
            Genres_list = soup.find_all('li', attrs={'class': 'versatility_linklist_item'})

            for Genres_item in Genres_list:
                Genres_text = Genres_item.text.replace('\n', '').strip()

                # やりたい出力1: ジャンル名を取得
                Genres_name = Genres_text.split('(')[0].strip()

                # やりたい出力2: 数値を取得
                match = re.search(r'\(([\d,]+)\)', Genres_text)
                if match:
                    Genres_num = int(match.group(1).replace(',', ''))
                else:
                    Genres_num = None

                details = {}
                datum = details
                datum['ジャンル'] = Genres_name
                datum['登録数'] = Genres_num
                data.append(datum)
                
            df = pd.DataFrame(data)
            df.index = np.arange(1, len(df)+1)
            df.to_excel(writer, sheet_name,index=False)

        else:
            pass           

print(f"Excelファイルに書き込みました: {filename}")
browser.quit()

#Excelファイル操作
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Excelファイルを開く
workbook = load_workbook(filename=filename)

# シート名が'ジャンル一覧(タグ)'である場合の処理
if 'ジャンル一覧(タグ)' in workbook.sheetnames:
    sheet = workbook['ジャンル一覧(タグ)']

    # 1行目に空欄行を挿入する
    sheet.insert_rows(0)
    # B列目に空欄列を挿入する
    sheet.insert_cols(1)

    # ソートする行を取得（ヘッダーを除く）
    rows_to_sort = list(sheet.iter_rows(min_row=3, max_row=sheet.max_row, values_only=True))

    # C列の値を整数として解釈し、数値の大きい順にソート
    sorted_rows = sorted(rows_to_sort, key=lambda row: int(row[2]) if row[2] is not None and str(row[2]).isdigit() else 0, reverse=True)

    # ソートされた行をシートに書き戻す（ヘッダーは変更しない）
    for row_idx, row_values in enumerate(sorted_rows, 3):
        for col_idx, value in enumerate(row_values, 1):
            sheet.cell(row=row_idx, column=col_idx, value=value)

    # 2行目の背景を灰色にする
    grey_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    for cell in sheet['2:2']:
        cell.fill = grey_fill

    # 列の幅を広げる (1単位 = 0.1956cm)
    sheet.column_dimensions['A'].width = 5.11
    sheet.column_dimensions['B'].width = 24.1

    # 罫線のスタイルを定義
    thin_border = Border(left=Side(style='thin'), 
                        right=Side(style='thin'), 
                        top=Side(style='thin'), 
                        bottom=Side(style='thin'))

    # B2からC310までのセルに細い罫線を引く
    for row in sheet.iter_rows(min_row=2, min_col=2, max_row=310, max_col=3):
        for cell in row:
            cell.border = thin_border

# 'genre'以外のシートに対する処理
for sheet_name in workbook.sheetnames:
    if sheet_name != 'ジャンル一覧(タグ)':
        sheet = workbook[sheet_name]
        # 1行目に空欄行を挿入する
        sheet.insert_rows(0)
        # B列目に空欄列を挿入する
        sheet.insert_cols(1)

        # 2行目の背景を灰色にする
        grey_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        for cell in sheet['2:2']:
            cell.fill = grey_fill

        # 列の幅を広げる (1単位 = 0.1956cm)
        sheet.column_dimensions['A'].width = 0.6 * 7.2
        sheet.column_dimensions['B'].width = 0.6 * 7.2
        sheet.column_dimensions['C'].width = 53.32
        sheet.column_dimensions['D'].width = 27.7
        sheet.column_dimensions['E'].width = 20.45
        sheet.column_dimensions['F'].width = 8.07
        sheet.column_dimensions['G'].width = 9.45
        sheet.column_dimensions['H'].width = 10.99
        sheet.column_dimensions['I'].width = 13.3
        sheet.column_dimensions['J'].width = 13.3
        sheet.column_dimensions['K'].width = 9.05
        sheet.column_dimensions['L'].width = 11.25
        sheet.column_dimensions['M'].width = 43
        sheet.column_dimensions['N'].width = 44.42
        sheet.column_dimensions['O'].width = 7.9244
        sheet.column_dimensions['P'].width = 23.1
        sheet.column_dimensions['Q'].width = 65.9

        # 細い罫線のスタイルを定義
        thin_border = Border(left=Side(style='thin'), 
                            right=Side(style='thin'), 
                            top=Side(style='thin'), 
                            bottom=Side(style='thin'))

        # B3からO102までのセルに細い罫線を引く
        for row in sheet.iter_rows(min_row=3, min_col=2, max_row=102, max_col=17):
            for cell in row:
                cell.border = thin_border

    else:
        pass

else:
    pass

# 変更を保存する
workbook.save(filename)

print(f"Excelファイルの見栄えを整えました: {filename}")