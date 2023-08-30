#ライブラリ呼び出し
from selenium import webdriver #ウェブサイトを自動で操作するためのツール「selenium」から、「webdriver」を取り入れる
from selenium.webdriver.chrome.options import Options
from selenium.common.exceptions import NoSuchElementException
from webdriver_manager.chrome import ChromeDriverManager #ChromeDriverを自動でダウンロードしてくれるツール「webdriver_manager」から、「ChromeDriverManager」を取り入れる
from bs4 import BeautifulSoup, Tag  #ウェブページの内容を解析するためのツール「BeautifulSoup」を取り入れています。
from time import sleep #「sleep」は、プログラムの実行を一時停止するためのツールです。
from datetime import datetime #「datetime」は、日付や時間を扱うためのツールです。
import os #「os」は、ファイルやディレクトリを操作するためのツールです。
import re
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import matplotlib.pyplot as plt
import matplotlib as mpl
import openpyxl as px #Excelファイルを操作するためのツール「openpyxl」を取り入れる
from openpyxl.styles import Border, Side
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.styles import Alignment
from openpyxl.styles import NamedStyle, Font, colors

# Chromeオプションの設定
chrome_options = Options()

# ヘッドレスモードを有効にする
chrome_options.add_argument("--headless")
chrome_options.add_argument('--disable-blinkfeatures=AutomationControlled')

# 画像を読み込まない設定を追加
prefs = {"profile.managed_default_content_settings.images": 2}
chrome_options.add_experimental_option("prefs", prefs)

# ChromeDriverManagerを使用してChromeドライバーを起動する
browser = webdriver.Chrome(ChromeDriverManager().install(), options=chrome_options)

# 出力するExcelファイル名
now_1 = datetime.now()
timestamp_str_1 = now_1.strftime("%Y%m%d%H%M")
filename = f'DLsite_同人ゲームランキング_{timestamp_str_1}.xlsx'

# Excelファイルに書き込むためのライターを作成
with pd.ExcelWriter(filename, engine='openpyxl') as writer:
    # 巡回したいURLと対応するシート名
    url_sheet_mapping = {
        #'https://www.dlsite.com/maniax/fsr/=/work_category[0]/doujin/order/trend/work_type_category[0]/game/options[0]/JPN/options[1]/NM/per_page/100/from/left_pain.work_type': '同人ゲームランキング1～100位'
        #'https://onl.sc/ECJz2XG':'101～200位'
        #'https://onl.sc/fMrqkDV': '201～300位'
        #'https://onl.sc/m99LWtf': '301～400位'
        #'https://onl.sc/tr74bBQ': '401～500位'
        'https://onl.sc/GHSamNv': '501～600位'
        #'https://onl.sc/HAm3Nbz': '601～700位'
        #'https://onl.sc/Pd8ksdw': '701～800位'
        #'https://onl.sc/8PAuAmd': '801～900位'
        #'https://onl.sc/TZg3Dzm': '901～1000位'
    }

    for url, sheet_name in url_sheet_mapping.items():   
        browser.get(url) # 指定したURLのページを開く
        sleep(1)
        try:
            elem_login_btn_top = browser.find_element_by_css_selector('.btn_yes.btn-approval') # ログインボタンの要素を指定する
            elem_login_btn_top.click() # ログインボタンをクリックする
        except NoSuchElementException:
            print('アクセス成功。'+str(sheet_name)+'Webページから全体データ取得中...')

        print('アクセス成功。'+str(sheet_name)+'Webページから全体データ取得中...')
        html_source = browser.page_source # ブラウザのページソースを取得する
        soup = BeautifulSoup(html_source, 'html.parser') # BeautifulSoupでブラウザのページソースを解析する
        
        data=[] 
        soup_shop = soup.find_all('dl', class_='work_img_main') # ページソースから、商品情報が記載されている部分を抽出する
        
        url_list = []
        for item in soup_shop:
            # 'a'タグ内の'href'属性を取得する
            url = item.find('a')['href']
            url_list.append(url)

        for index,item in enumerate(url_list):
            index_plus = index + 501 #順位が違う場合変える
            print(str(sheet_name)+f'No.{index_plus}のURLを取得中...')

            dig_url = item

            # Chromeオプションの設定
            chrome_options = Options()
            chrome_options.add_argument("--headless")             # ヘッドレスモードを有効にする
            prefs = {"profile.managed_default_content_settings.images": 2}             # 画像を読み込まない設定を追加
            chrome_options.add_argument('--disable-blinkfeatures=AutomationControlled')
            chrome_options.add_experimental_option("prefs", prefs)
            # ChromeDriverManagerを使用してChromeドライバーを起動する
            browser = webdriver.Chrome(ChromeDriverManager().install(), options=chrome_options)

            browser.get(dig_url) # 指定したURLのページを開く

            try:
                elem_login_btn_top = browser.find_element_by_css_selector('.btn_yes.btn-approval') # ログインボタンの要素を指定する
                elem_login_btn_top.click() # ログインボタンをクリックする
            except NoSuchElementException:
                pass

            print('アクセス成功。'+str(sheet_name)+str(index_plus )+'位の個別Webページのデータ取得中')
            html_source = browser.page_source # ブラウザのページソースを取得する
            shop_page_soup = BeautifulSoup(html_source, 'html.parser') # BeautifulSoupでブラウザのページソースを解析する

            print(str(sheet_name)+str(index_plus )+'位の個別Webページのデータ解析中')

            #24時間、週間、月間ランキング情報取得
            ranking_items_elem = shop_page_soup.find('dl', 'work_right_info_title')
            ranking_items = ranking_items_elem.find_all('li')

            #24時間ランキング
            try:
                ranking_items_day = int(ranking_items[0].text.replace('位', '').replace('24時間', ''))
            except Exception as e:
                ranking_items_day = '―'  # 空欄表記

            #週間ランキング
            try:
                ranking_items_week = int(ranking_items[1].text.replace('位', '').replace('7日間', ''))
            except Exception as e:
                ranking_items_week =  '―'  # 空欄表記

            #月間ランキング
            try:
                ranking_items_month = int(ranking_items[2].text.replace('位', '').replace('30日間', ''))
            except Exception as e:
                ranking_items_month =  '―'  # 空欄表記

            #専売の有無
            dlsite_exclusive = "×"
            try:
                exclusive_items = shop_page_soup.find_all('span', class_=lambda x: x and x.startswith('icon_lead_01'))
                if "DLsite専売" in exclusive_items[0].text:
                    dlsite_exclusive = "○"
                else:
                    pass
            except Exception as e:
                pass    

            #商品名
            product_title = shop_page_soup.find('h1', {'id': 'work_name', 'itemprop': 'name'})
            try:
                product_title_str = product_title.text.strip()
            except Exception as e:
                pass

            print('商品名取得。'+str(sheet_name)+str(index_plus )+'位の商品名は、「' +product_title_str+'」です。')
            
            #新作・旧作表記
            element = shop_page_soup.find('span', 'c_icon_latest')
            if element is not None:
                latest_situation = element.text.strip()
            else:
                latest_situation = '―'

            #ブランド名の取得
            brand_name_element = shop_page_soup.find('span', {'class': 'maker_name', 'itemprop': 'brand'})
            brand_name_text = brand_name_element.text.strip().replace(",", "")

            #ブランドURLの取得
            try:    
                brand_url_element = shop_page_soup.find('span', {'class': 'maker_name', 'itemprop': 'brand'})
                if brand_url_element:
                    brand_url_item = brand_url_element.find('a')
                    if brand_url_item:
                        brand_url = brand_url_item['href']
            except Exception as e:
                pass

            #作品形式タグの取得
            item_genre_name_elements = shop_page_soup.find_all('div', {'class': 'work_genre', 'id': 'category_type'})
            for element in item_genre_name_elements:
                if element.text:
                    item_genre_name_text = element.text.strip()
                    break
            else:
                item_genre_name_text  = '―'  # 空欄表記

            #トータル販売数の取得
            try:
                number_of_sales = shop_page_soup.find('dd', class_='point')
                number_of_sales_int = int(number_of_sales.text.strip().replace(",", ""))
            except Exception as e:  
                dd_tag = shop_page_soup.find('dd', class_='work_dl_details point')
                extracted_text = dd_tag.text.strip()
                number_of_sales= extracted_text.split()[0].replace(",", "")

            #販売価格(割引価格)
            price_item = shop_page_soup.find('div', class_='work_buy_container')
            selling_price = price_item.find('strong', class_=lambda x: x and x.startswith('price')) 
            if  selling_price  is not None:
                selling_price_str = selling_price.text.strip().replace('円', '')
                selling_price_int = int(re.search(r'(\d+)',selling_price_str.replace(',', '')).group(1))
            else:
                selling_price_int  = '―'  # 空欄表記
   
            #割引率
            discount_rate = shop_page_soup.find('p', {'class': 'type_sale', 'data-toggle': 'found'})
            if  discount_rate is not None:
                discount_rate_str = discount_rate.text.strip().replace('円', '')
                discount_rate_int = int(re.search(r'(\d+)',discount_rate_str.replace(',', '')).group(1))
            else:
                discount_rate_int  = '―'  # 空欄表記

            #サークル設定価格
            main_price = price_item.find('span', class_=lambda x: x and x.startswith('strike'))
            if  main_price is not None:
                main_price_str = main_price.text.strip().replace('円', '')
                main_price_int = int(re.search(r'(\d+)',main_price_str.replace(',', '')).group(1))
            else:
                main_price = shop_page_soup.find('span', {'class': 'price'})
                main_price_str = main_price.text.strip().replace('円', '')
                main_price_int = int(re.search(r'(\d+)',main_price_str.replace(',', '')).group(1))

            #卸売価格ベースの販売予想額
            if selling_price_int == '―' :
                if main_price_int < 1000:
                    ratio = 0.5
                elif main_price_int < 2000:
                    ratio = 0.6364
                elif main_price_int< 3000:
                    ratio =  0.7143
                elif main_price_int < 4000:
                    ratio =  0.7647
                else:
                    ratio =  0.8  # 販売価格4000円以上の場合、卸価格は販売価格の80%
                # 卸売価格を計算
                wholesale_price = int(main_price_int * ratio)
                wholesale_price_result = wholesale_price * number_of_sales_int #販売予想額 
            else:
                def calculate_wholesale_price(selling_price_int):
                    if selling_price_int < 1000:
                        ratio = 0.5
                    elif selling_price_int < 2000:
                        ratio = 0.6364
                    elif selling_price_int < 3000:
                        ratio =  0.7143
                    elif selling_price_int < 4000:
                        ratio =  0.7647
                    else:
                        ratio =  0.8  # 販売価格4000円以上の場合、卸価格は販売価格の80%
                    # 卸売価格を計算
                    wholesale_price = selling_price_int * ratio
                    return int(wholesale_price)  # 結果を整数として返す
                wholesale_price_result = calculate_wholesale_price(selling_price_int) * number_of_sales_int #販売予想額

            #性癖ジャンルタグの取得
            genre_name_elements = shop_page_soup.find_all('div', class_='main_genre')
            for element in genre_name_elements:
                if element.text:
                    genre_name_text = element.text.strip().replace("\n", "／")
                    break
            else:
                genre_name_text  = '―'  # 空欄表記

            #平均評価
            rewiew_average_item = shop_page_soup.find('span',class_='point average_count')
            if rewiew_average_item is not None:
                rewiew_average = rewiew_average_item.text
            else:
                rewiew_average = '―'

            #総評価数
            rewiew_evaluate_item = shop_page_soup.find('dd','star_box')
            if rewiew_evaluate_item is not None:
                rewiew_evaluate = rewiew_evaluate_item.find('span','count').text.replace(",", "").replace("(", "").replace(")", "")
            else:
                rewiew_evaluate = '―'

            #レビュー数
            try:
                rewiew_num_item = shop_page_soup.find_all('dd','position_fix')[1]
                rewiew_num = int(re.sub(r'[^\d,]', '', rewiew_num_item.text).replace(',', ''))
            except Exception as e:  
                rewiew_num = '―'

            #お気に入り登録数
            try:
                favorite_evaluate_item = shop_page_soup.find_all('dd','position_fix')[0]
                favorite_number = int(re.sub(r'[^\d,]', '', favorite_evaluate_item.text).replace(',', ''))
            except Exception as e: 
                favorite_number = '―'

            #「まとめ記事」登録数
            try:
                fan_evaluate_item = shop_page_soup.select_one('.dl_ch_button.matome')
                fan_evaluate = fan_evaluate_item.find('span','count')
                fan_evaluate_num = int(fan_evaluate.text.replace('件' , ''))
            except TypeError:   # select_oneがNoneを返す場合、TypeErrorが発生する
                try:    # 最初の処理がダメな場合、2つ目の処理を試みる
                    fan_evaluate_item = shop_page_soup.find('div',class_='link_dl_ch')
                    fan_evaluate = fan_evaluate_item.find('span','count')
                    fan_evaluate_num = int(fan_evaluate.text.replace('件' , ''))
                except TypeError:  # findがNoneを返す場合、TypeErrorが発生する
                    fan_evaluate_num = '―'
                except Exception as e: 
                    fan_evaluate_num = '―'
            except Exception as e: 
                fan_evaluate_num = '―'
            
            #まとめ記事のURL
            try:
                fan_evaluate_item = shop_page_soup.select_one('.dl_ch_button.matome')
                href_value = fan_evaluate_item['href']
                fan_evaluate_URL = 'https:' + href_value
            except TypeError:  
                try:
                    fan_evaluate_item = shop_page_soup.find('div', class_='link_dl_ch')
                    href_value = fan_evaluate_item.find('a')['href']
                    fan_evaluate_URL = 'https:' + href_value
                except TypeError:  # findがNoneを返す場合、TypeErrorが発生する
                    fan_evaluate_URL = '―'
                except Exception as e:
                    fan_evaluate_URL = '―'
            except Exception as e:
                fan_evaluate_URL = '―'

            #配信開始日
            try:
                sells_day_item = shop_page_soup.find('table', {'id': 'work_outline', 'cellspacing': '0'})
                sells_day_item_num = sells_day_item.find_all('td')[0].text.replace(' ', '')

                def reformat_date(input_str):
                    # 正規表現で年、月、日を検出
                    match = re.search(r'(\d{4})年\s*(\d{1,2})月\s*(\d{1,2})日?', input_str)
                    
                    if match:
                        year = match.group(1)
                        month = match.group(2)
                        day = match.group(3)                       
                        # 日付を"YYYY/M/D"の形式に整形
                        formatted_date = f"{year}/{int(month)}/{int(day)}"                       
                        return formatted_date
                    else:
                        pass               
                
                target_date = reformat_date(sells_day_item_num) 
            
            except Exception as e: 
                target_date = '―'

            # 配信後の日数
            now_2 = datetime.now()
            timestamp_2 = now_2.strftime("%Y%m%d")
            date_format_target = "%Y/%m/%d"  # target_dateの形式を変更
            date_format_timestamp = "%Y%m%d"
            target_date_obj = datetime.strptime(target_date, date_format_target)
            timestamp_2_obj = datetime.strptime(timestamp_2, date_format_timestamp)
            pass_day = (timestamp_2_obj - target_date_obj).days

            #ファイル容量
            try:
                file_item = shop_page_soup.find('table', {'id': 'work_outline', 'cellspacing': '0'})
                if 'ファイル容量' in file_item.text:
                    file_vol_element = file_item.find_all('div','main_genre')[1]
                    file_vol = file_vol_element.text.strip().replace("\n", "")
                else:
                    file_vol = '―'  # 空欄表記
            except Exception as e: 
                file_vol = '―'

            # details辞書にデータを格納
            details = {}
            datum=details
            details['No'] = (index_plus)
            details['24\n時間\nﾗﾝ\nｷﾝｸﾞ'] = ranking_items_day
            details['週間\nﾗﾝ\nｷﾝｸﾞ'] = ranking_items_week
            details['月間\nﾗﾝ\nｷﾝｸﾞ'] = ranking_items_month
            details['DL\nｻｲﾄ\n専売'] = dlsite_exclusive
            details['商品名'] = product_title_str
            details['新旧\n表記'] = latest_situation
            details['ﾌﾞﾗﾝﾄﾞ名'] = brand_name_text 
            details['ゲーム\nｼﾞｬﾝﾙ'] = item_genre_name_text
            details['ﾄｰﾀﾙ\n販売数'] =  number_of_sales_int
            details['販売\n価格'] = selling_price_int
            details['登録\n売価'] = main_price_int 
            details['割\n引\n率'] = discount_rate_int
            details['予想\n粗利\n額'] = wholesale_price_result 
            details['ｼﾞｬﾝﾙ\nタグ'] = genre_name_text
            details['平均\n評価'] = rewiew_average
            details['総評\n価数'] = rewiew_evaluate
            details['お気に\n入り\n登録数'] = favorite_number 
            details['まとめ\n記事\n登録数'] = fan_evaluate_num
            details['配信\n開始日'] = target_date
            details['本ﾃﾞｰﾀ\n作成日\nとの差\n(日数)'] = pass_day
            details['ﾌｧｲﾙ\n容量'] = file_vol 
            details['商品\nURL'] = dig_url
            details['ブランド\nURL'] = brand_url
            details['まとめ\n記事\nURL'] = fan_evaluate_URL
            data.append(datum)

            print(str(sheet_name)+str(index_plus )+'位の個別データ書き込み完了……')
            browser.quit()
            sleep(1) 

        df = pd.DataFrame(data)
        df.index = np.arange(1, len(df)+1)
        df.to_excel(writer, sheet_name,index=False)

        # ユーザーのデスクトップパスを取得
        desktop_path = os.path.join(os.path.join(os.environ['USERPROFILE']), 'Desktop')
        # 今日の日付を取得
        today_date = datetime.today().strftime('%Y%m%d')
        # 保存先のフォルダ名を指定（今日の日付を追加）
        folder_name = 'img_' + today_date
        # フルパスを作成
        save_path = os.path.join(desktop_path, folder_name)
        # フォルダが存在しない場合、作成
        if not os.path.exists(save_path):
            os.makedirs(save_path)
            
        # 日本語フォントの設定
        mpl.rcParams['font.family'] = 'Yu Mincho' # 例として「Yu Mincho」を使用
        plt.rcParams['font.family'] = 'Yu Gothic' # もしくは他の日本語対応のフォント名
        
        # '－'をdf['登録\n売価']と同じ値にする
        mask = df['販売\n価格'] == '―'
        df.loc[mask, '販売\n価格'] = df.loc[mask, '登録\n売価']

        # 販売数と現状価格のプロット
        plt.figure()
        plt.scatter(df['販売\n価格'],df['ﾄｰﾀﾙ\n販売数'])  # 点をプロット
        plt.title('販売価格と販売数', fontsize=18, fontweight='bold')  # タイトルをゴシック、太字、18ポイントに
        plt.xlabel('販売価格')  # x軸のラベル
        plt.ylabel('販売数')  # y軸のラベル
        plt.savefig(save_path + '\\graph_' + str(sheet_name) + '1.png')
        plt.close()

        plt.figure()
        plt.scatter(df['登録\n売価'].astype(float), df['販売\n価格'].astype(float))  # 点をプロット
        plt.title('登録売価と販売価格', fontsize=18, fontweight='bold')
        plt.xlabel('登録売価')  # x軸のラベル
        plt.ylabel('販売価格')  # y軸のラベル
        plt.savefig(save_path + '\\graph_' + str(sheet_name) + '2.png')
        plt.close()

        # サークル数と粗利金額のローレンツ曲線のプロット
        plt.figure()
        plt.scatter(range(len(df['予想\n粗利\n額'])), [x / 10000 for x in sorted(df['予想\n粗利\n額'])])  # 点をプロット、y値を10000で割る
        plt.title('サークル数(累計数)と粗利金額', fontsize=18, fontweight='bold')
        plt.xlabel('サークル数(累計数)')  # x軸のラベル
        plt.ylabel('粗利金額 (万)')    # y軸のラベルに「万」を追加
        plt.ticklabel_format(style='plain', axis='y')  # y軸のラベルを指数表記にしない
        plt.savefig(save_path + '\\graph_' + str(sheet_name) + '3.png')
        plt.close()

        plt.figure()
        plt.scatter(df['本ﾃﾞｰﾀ\n作成日\nとの差\n(日数)'], [x / 10000 for x in sorted(df['予想\n粗利\n額'])])  # 点をプロット
        plt.title('経過日数と粗利金額', fontsize=18, fontweight='bold')
        plt.xlabel('経過日数')  # x軸のラベル
        plt.ylabel('粗利金額 (万)')  # y軸のラベル
        plt.savefig(save_path + '\\graph_' + str(sheet_name) + '4.png')
        plt.close()
        
        browser.quit()

print(f"Excelファイルに書き込み完了……: {filename}")
browser.quit()

print(f"Excelファイルの見栄えを整えます……: {filename}")
#Excelファイル操作
workbook = load_workbook(filename=filename) # Excelファイルを開く
for sheet_name in workbook.sheetnames:
    sheet = workbook[sheet_name]
    # 1行目に空欄行を挿入する
    sheet.insert_rows(0)
    # B列目に空欄列を挿入する
    sheet.insert_cols(1)

    # 2行目の背景を灰色にする
    grey_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    for cell in sheet['2:2']:
        cell.fill = grey_fill

    # 列の幅を広げる (1単位 = 0.1956cm)
    sheet.column_dimensions['A'].width = 0.52 / 0.1956
    sheet.column_dimensions['B'].width = 1.01 / 0.1956
    sheet.column_dimensions['C'].width = 1.01 / 0.1956
    sheet.column_dimensions['D'].width = 1.01 / 0.1956
    sheet.column_dimensions['E'].width = 1.01 / 0.1956
    sheet.column_dimensions['F'].width = 1.01 / 0.1956
    sheet.column_dimensions['G'].width = 6.40 / 0.1956
    sheet.column_dimensions['H'].width = 1.41 / 0.1956
    sheet.column_dimensions['I'].width = 4.6 / 0.1956
    sheet.column_dimensions['J'].width = 3.92 / 0.1956
    sheet.column_dimensions['K'].width = 1.39 / 0.1956
    sheet.column_dimensions['L'].width = 1.39 / 0.1956  
    sheet.column_dimensions['M'].width = 1.39 / 0.1956
    sheet.column_dimensions['N'].width = 1.12 / 0.1956
    sheet.column_dimensions['O'].width = 2.23 / 0.1956
    sheet.column_dimensions['P'].width = 8.20 / 0.1956
    sheet.column_dimensions['Q'].width = 1.06 / 0.1956
    sheet.column_dimensions['R'].width = 1.06 / 0.1956
    sheet.column_dimensions['S'].width = 1.61 / 0.1956
    sheet.column_dimensions['T'].width = 1.36 / 0.1956
    sheet.column_dimensions['U'].width = 1.47 / 0.1956
    sheet.column_dimensions['V'].width = 1.44 / 0.1956
    sheet.column_dimensions['W'].width = 1.6 / 0.1956
    sheet.column_dimensions['X'].width = 2.20 / 0.1956
    sheet.column_dimensions['Y'].width = 2.20 / 0.1956
    sheet.column_dimensions['Z'].width = 2.20 / 0.195

    # 行の幅を広げる
    sheet.row_dimensions[2].height = 12 / 0.1956 

    # 3～102行目の全てのセルに対して、中央揃えを適用
    for row in sheet.iter_rows(min_row=3, max_row=102):
        for cell in row:
            cell.alignment = Alignment(horizontal='center') # 中央揃えのスタイルを適用

    #左揃えに設定
    for col in ['G','H','I','J','P','X','Y','Z']:
        for cell in sheet[col]:
            cell.alignment = Alignment(horizontal='left') # 左揃えのスタイルを適用
    
    #右揃えに設定
    for col in ['K','L','M','N','O','Q','R','S','T','U','V','W']:
        for cell in sheet[col]:
            cell.alignment = Alignment(horizontal='right') # 右揃えのスタイルを適用

    # 2行目の各セルに対して、縦・横方向の中央揃え
    for cell in sheet['2:2']:
        cell.alignment = Alignment(vertical='center')
        cell.alignment = Alignment(horizontal='center') 

    # 細い罫線のスタイルを定義
    thin_border = Border(left=Side(style='thin'), 
                        right=Side(style='thin'), 
                        top=Side(style='thin'), 
                        bottom=Side(style='thin'))

    # B3からO102までのセルに細い罫線を引く
    for row in sheet.iter_rows(min_row=3, min_col=2, max_row=102, max_col=26):
        for cell in row:
            cell.border = thin_border

    # ハイパーリンク用のスタイルを定義
    for idx, url in enumerate(df['商品\nURL'], start=3):  
        sheet.cell(row=idx, column=24, value=url).hyperlink = url 

    for idx, url in enumerate(df['ブランド\nURL'], start=3):
        sheet.cell(row=idx, column=25, value=url).hyperlink = url  

    for idx, url in enumerate(df['まとめ\n記事\nURL'], start=3): 
        sheet.cell(row=idx, column=26, value=url).hyperlink = url  

    # 画像を挿入
    img1 = px.drawing.image.Image(save_path +'\graph_' + sheet_name + '1.png')
    img1.anchor = 'B104'
    sheet.add_image(img1)

    img2 = px.drawing.image.Image(save_path +'\graph_' + sheet_name + '2.png')
    img2.anchor = 'J104' # 位置を調整する必要があります
    sheet.add_image(img2)

    img3 = px.drawing.image.Image(save_path +'\graph_' + sheet_name + '3.png')
    img3.anchor = 'B129' # 位置を調整する必要があります
    sheet.add_image(img3)

    img4 = px.drawing.image.Image(save_path +'\graph_' + sheet_name + '4.png')
    img4.anchor = 'J129' # 位置を調整する必要があります
    sheet.add_image(img4)

# 変更を保存する
workbook.save(filename)

print(f"Excelファイルの見栄えを整えました。全工程完了: {filename}")