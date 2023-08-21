from selenium import webdriver #ウェブサイトを自動で操作するためのツール「selenium」から、「webdriver」を取り入れる
from selenium.webdriver.chrome.options import Options
from bs4 import BeautifulSoup #ブページの内容を解析するためのツール「BeautifulSoup」を取り入れています。
from time import sleep #「sleep」は、プログラムの実行を一時停止するためのツールです。
from webdriver_manager.chrome import ChromeDriverManager #ChromeDriverを自動でダウンロードしてくれるツール「webdriver_manager」から、「ChromeDriverManager」を取り入れる
from datetime import datetime #「datetime」は、日付や時間を扱うためのツールです。
import re
import pandas as pd
import numpy as np 
import openpyxl #Excelファイルを操作するためのツール「openpyxl」を取り入れる
from openpyxl.styles import Font #「Font」は、Excelファイルのセルの書式を操作するためのツールです。
from selenium.common.exceptions import NoSuchElementException


# ヘッドレスモードのオプションを設定する
chrome_options = Options()
chrome_options.add_argument("--headless")

# ChromeDriverManagerを使用してChromeドライバーを起動する
browser = webdriver.Chrome(ChromeDriverManager().install(), options=chrome_options)

# 出力するExcelファイル名
now = datetime.now()
timestamp_str = now.strftime("%Y%m%d%H%M%S")
filename = f'dlsite_sales_ranking_data_{timestamp_str}.xlsx'

# Excelファイルに書き込むためのライターを作成
with pd.ExcelWriter(filename, engine='openpyxl') as writer:
    # 巡回したいURLと対応するシート名
    url_sheet_mapping = {
        'https://www.dlsite.com/maniax/ranking/day': 'day',
        'https://www.dlsite.com/maniax/ranking/week': 'week',
        'https://www.dlsite.com/maniax/ranking/month': 'month',
        'https://www.dlsite.com/maniax/ranking/year': 'year',
        'https://www.dlsite.com/maniax/ranking/total': 'total',
        'https://www.dlsite.com/maniax/genre/list': 'genre'
    }

    for url, sheet_name in url_sheet_mapping.items():   
        browser.get(url) # 指定したURLのページを開く
        sleep(1) # 1秒間待機する

        if sheet_name != 'genre':
            try:
                elem_login_btn_top = browser.find_element_by_css_selector('.btn_yes.btn-approval') # ログインボタンの要素を指定する
                elem_login_btn_top.click() # ログインボタンをクリックする
            except NoSuchElementException:
                print("R18以上チェックボタン確認不可。次の処理へ進みます。")

            html_source = browser.page_source # ブラウザのページソースを取得する
            soup = BeautifulSoup(html_source, 'html.parser') # BeautifulSoupでブラウザのページソースを解析する
            
            #商品ページを1つに絞る
            data=[]
            products = soup.select('tr')
            for product in products:
                #ランキングナンバーの取得
                rank_elements = product.find('div', class_=lambda x: x and x.startswith('rank_no'))
                rank_no = int(rank_elements.text)  # 最初の要素のテキストを整数に変換

                #タイトルの取得
                title=product.select('a')[1]
                title=title.text.strip()

                #サークル名の取得
                circul_name=product.select('a')[2]
                circul_name=circul_name.text.strip()

                #製品ジャンルの取得
                category_name=product.find('div',attrs = {'class':'work_category'})
                category_name=category_name.text.strip()

                #販売数の取得
                dl_count_div = product.find('div', class_='dl_count')
                dl_count_text = dl_count_div.text
                dl_count = int(''.join(filter(str.isdigit, dl_count_text)))
                #filter 関数と str.isdigit メソッドを使って、文字列から数字だけを取り出します。
                #join メソッドを使って、取り出した数字を文字列に結合します。
                #int 関数を使って、文字列を整数型に変換します。

                #販売価格の取得
                work_price = product.find('span', class_=lambda x: x and x.startswith('work_price'))
                work_price_str = work_price.text.replace(',', '').replace('円', '') # テキスト内容を取得し、カンマを削除
                work_price_int = int(work_price_str) # 文字列を整数型に変換

                #登録売価の取得
                strike_price = product.find('span', attrs={'class': 'strike'})
                if strike_price is not None:
                    strike_price_str = strike_price.text.replace(',', '').replace('円', '')
                    strike_price_int = int(strike_price_str)
                else:
                    strike_price_int = '※登録売価からの割引はない'  

                #割引率の取得
                discount_percentage=product.find('span',attrs = {'class':'icon_campaign'}) # テキスト内容を取得し、%の前の数字部分だけを取り出す
                if discount_percentage is not None:
                    discount_percentage_str = discount_percentage.text.split('%')[0]
                    discount_percentage_int = int(discount_percentage_str)
                else:
                    discount_percentage_int= ''  # 空欄表記

                #割引終了時期の取得
                priod_date = product.find('span', attrs={'class': 'period_date'}) # テキスト内容を取得
                if priod_date is not None:
                    priod_date_text = priod_date.text
                    date_match = re.search(r'(\d{4})年(\d{2})月(\d{2})日 (\d{2})時(\d{2})分', priod_date_text)
                    #正規表現を使って、日付と時刻のパターンに一致する部分を検索します。この例では、正規表現の (\d{4})年(\d{2})月(\d{2})日 (\d{2})時(\d{2})分 が使用されています。
                    formatted_date = f'{date_match.group(1)}/{int(date_match.group(2))}/{date_match.group(3)}/{date_match.group(4)}:{date_match.group(5)}'
                    #re.search の結果から、各部分（年、月、日、時、分）を取り出し、所望のフォーマットに整形します。
                else:
                    formatted_date = ''  # 空欄表記

                #タグの取得
                search_tag = product.find('dd', class_='search_tag') # クラス名が"search_tag"であるddタグを探す
                if search_tag is not None:
                    tags = [a_tag.text for a_tag in search_tag.find_all('a')] # aタグのテキスト内容を取得し、一つの文字列に結合
                    output_string = '／'.join(tags) # タグを「／」で区切って一つの文字列に結合
                else:
                    output_string = ''  # 空欄表記

                #紹介文の取得
                work_text = product.find('dd', attrs={'class': 'work_text'}) 
                work_text=work_text.text

                #紹介文の文字数取得
                work_text_len=len(work_text)

                #「音声」「音楽」「体験版」などの有無
                work_genre_tag = product.find('dd', class_='work_genre') # クラス名が"work_genre"であるddタグを探す
                attributes = [span_tag.text for span_tag in work_genre_tag.find_all('span') if span_tag.text] # spanタグのテキスト内容を取得し、一つの文字列に結合
                output_string = '／'.join(attributes)

                #URL
                product_url_tag = product.find('a', class_='work_thumb_box')
                product_url = product_url_tag['href'] # href属性を取得

                details = {}
                datum=details
                datum['No.']=rank_no
                datum['商品名']=title
                datum['ｻｰｸﾙ名']=circul_name
                datum['ｼﾞｬﾝﾙ']=category_name
                datum['販売数']=dl_count 
                datum['現状価格']=work_price_int  
                datum['登録売価']=strike_price_int 
                datum['割引率']=discount_percentage_int
                datum['割引終了']=formatted_date
                datum['タグ']=output_string 
                datum['紹介文']=work_text
                datum['紹介文文字数']=work_text_len
                datum['音声等']=output_string 
                datum['URL']=product_url
                data.append(datum)

            df = pd.DataFrame(data)
            df.index = np.arange(1, len(df)+1)
            df.to_excel(writer, sheet_name,index=False)

        elif sheet_name == 'genre':
            try:
                elem_login_btn_top = browser.find_element_by_css_selector('.btn_yes.btn-approval') # ログインボタンの要素を指定する
                elem_login_btn_top.click() # ログインボタンをクリックする
            except NoSuchElementException:
                print("R18以上チェックボタン確認不可。次の処理へ進みます。")

            html_source = browser.page_source # ブラウザのページソースを取得する
            soup = BeautifulSoup(html_source, 'html.parser') # BeautifulSoupでブラウザのページソースを解析する
           
            data = []
            Genres_list = soup.find_all('li', attrs={'class': 'versatility_linklist_item'})

            for Genres_item in Genres_list:
                Genres_text = Genres_item.text.replace('\n', '').strip()

                # やりたい出力1: ジャンル名を取得
                Genres_name = Genres_text.split('(')[0].strip()

                # やりたい出力2: 数値を取得
                match = re.search(r'\(([\d,]+)\)', Genres_text)
                if match:
                    Genres_num = int(match.group(1).replace(',', ''))
                else:
                    Genres_num = None

                details = {}
                datum = details
                datum['ジャンル'] = Genres_name
                datum['登録数'] = Genres_num
                data.append(datum)
                
            df = pd.DataFrame(data)
            df.index = np.arange(1, len(df)+1)
            df.to_excel(writer, sheet_name,index=False)

        else:
            pass           

print(f"Excelファイルに書き込みました: {filename}")
browser.quit()

#Excelファイル操作
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Excelファイルを開く
workbook = load_workbook(filename=filename)

# シート名が'genre'である場合の処理
if 'genre' in workbook.sheetnames:
    sheet = workbook['genre']

    # 1行目に空欄行を挿入する
    sheet.insert_rows(0)
    # B列目に空欄列を挿入する
    sheet.insert_cols(1)

    # ソートする行を取得（ヘッダーを除く）
    rows_to_sort = list(sheet.iter_rows(min_row=3, max_row=sheet.max_row, values_only=True))

    # C列の値を整数として解釈し、数値の大きい順にソート
    sorted_rows = sorted(rows_to_sort, key=lambda row: int(row[2]) if row[2] is not None and str(row[2]).isdigit() else 0, reverse=True)

    # ソートされた行をシートに書き戻す（ヘッダーは変更しない）
    for row_idx, row_values in enumerate(sorted_rows, 3):
        for col_idx, value in enumerate(row_values, 1):
            sheet.cell(row=row_idx, column=col_idx, value=value)

    # 2行目の背景を灰色にする
    grey_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    for cell in sheet['2:2']:
        cell.fill = grey_fill

    # 列の幅を広げる (1インチ = 72ポイント)
    sheet.column_dimensions['A'].width = 1 * 7.2
    sheet.column_dimensions['B'].width = 6 * 7.2
    sheet.column_dimensions['A'].width = 2 * 7.2

# 'genre'以外のシートに対する処理
for sheet_name in workbook.sheetnames:
    if sheet_name != 'genre':
        sheet = workbook[sheet_name]
        # 1行目に空欄行を挿入する
        sheet.insert_rows(0)
        # B列目に空欄列を挿入する
        sheet.insert_cols(1)

        # 2行目の背景を灰色にする
        grey_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        for cell in sheet['2:2']:
            cell.fill = grey_fill

        # 列の幅を広げる (1インチ = 72ポイント)
        sheet.column_dimensions['A'].width = 0.6 * 7.2
        sheet.column_dimensions['B'].width = 0.6 * 7.2
        sheet.column_dimensions['C'].width = 6 * 7.2
        sheet.column_dimensions['D'].width = 2.5 * 7.2
        sheet.column_dimensions['E'].width = 1.6 * 7.2
        sheet.column_dimensions['F'].width = 1.2 * 7.2
        sheet.column_dimensions['G'].width = 1.2 * 7.2
        sheet.column_dimensions['H'].width = 1.2 * 7.2
        sheet.column_dimensions['I'].width = 0.6 * 7.2
        sheet.column_dimensions['J'].width = 1.8 * 7.2
        sheet.column_dimensions['K'].width = 1.5 * 7.2
        sheet.column_dimensions['L'].width = 5 * 7.2
        sheet.column_dimensions['M'].width = 1.2 * 7.2
        sheet.column_dimensions['N'].width = 2 * 7.2
        sheet.column_dimensions['O'].width = 8 * 7.2 
    else:
        pass

else:
    pass

# 変更を保存する
workbook.save(filename)

print(f"Excelファイルの見栄えを整えました: {filename}")