#ライブラリ呼び出し
from selenium import webdriver #ウェブサイトを自動で操作するためのツール「selenium」から、「webdriver」を取り入れる
from selenium.webdriver.chrome.options import Options
from selenium.common.exceptions import NoSuchElementException
from webdriver_manager.chrome import ChromeDriverManager #ChromeDriverを自動でダウンロードしてくれるツール「webdriver_manager」から、「ChromeDriverManager」を取り入れる
from bs4 import BeautifulSoup #ウェブページの内容を解析するためのツール「BeautifulSoup」を取り入れています。
from time import sleep #「sleep」は、プログラムの実行を一時停止するためのツールです。
from datetime import datetime #「datetime」は、日付や時間を扱うためのツールです。
import os #「os」は、ファイルやディレクトリを操作するためのツールです。
import re
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import matplotlib.pyplot as plt
import matplotlib as mpl
import openpyxl as px #Excelファイルを操作するためのツール「openpyxl」を取り入れる
from openpyxl.styles import Border, Side
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.styles import Alignment
from openpyxl.styles import NamedStyle, Font, colors

# 出力するExcelファイル名
now_1 = datetime.now()
timestamp_str_1 = now_1.strftime("%Y%m%d%H%M")
filename = f'FANZA_同人ランキング_{timestamp_str_1}.xlsx'

# Chromeオプションの設定
chrome_options = Options()
# ヘッドレスモードを有効にする
chrome_options.add_argument("--headless")
chrome_options.add_argument('--disable-blinkfeatures=AutomationControlled')
# 画像を読み込まない設定を追加
prefs = {"profile.managed_default_content_settings.images": 2}
chrome_options.add_experimental_option("prefs", prefs)
# ChromeDriverManagerを使用してChromeドライバーを起動する
browser = webdriver.Chrome(ChromeDriverManager().install(), options=chrome_options)

# Excelファイルに書き込むためのライターを作成
with pd.ExcelWriter(filename, engine='openpyxl') as writer:
    # 巡回したいURLと対応するシート名
    url_sheet_mapping = {
        'https://www.dmm.co.jp/dc/doujin/-/list/=/limit=120/media=game/section=mens/page=1/': '同人ゲーム1～120位'
        #'https://www.dmm.co.jp/dc/doujin/-/list/=/limit=120/media=game/section=mens/page=2/':'同人ゲーム121～240位'
        #'https://www.dmm.co.jp/dc/doujin/-/list/=/limit=120/media=game/section=mens/page=3/':'同人ゲーム241～360位'
        #'https://www.dmm.co.jp/dc/doujin/-/list/=/limit=120/media=game/section=mens/page=4/':'同人ゲーム361～480位'
        #'https://www.dmm.co.jp/dc/doujin/-/list/=/limit=120/media=game/section=mens/page=5/':'同人ゲーム481～600位'
        #'https://www.dmm.co.jp/dc/doujin/-/list/=/limit=120/media=game/section=mens/page=6/':'同人ゲーム601～720位'
        #'https://www.dmm.co.jp/dc/doujin/-/list/=/limit=120/media=game/section=mens/page=7/':'同人ゲーム721～840位'
        #'https://www.dmm.co.jp/dc/doujin/-/list/=/limit=120/media=game/section=mens/page=8/':'同人ゲーム841～960位'
        #'https://www.dmm.co.jp/dc/doujin/-/list/=/limit=120/media=game/section=mens/page=9/':'同人ゲーム961～1080位'
        #'https://www.dmm.co.jp/dc/doujin/-/list/=/limit=120/media=game/section=mens/page=10/':'同人ゲーム1081～1200位'    
    }
    for url, sheet_name in url_sheet_mapping.items():   
        browser.get(url) # 指定したURLのページを開く
        sleep(1)

        try:
            elem_login_btn_top = browser.find_element_by_css_selector('.ageCheck__link--r18') # ログインボタンの要素を指定する
            elem_login_btn_top.click() # ログインボタンをクリックする
        except NoSuchElementException:
            print('アクセス成功。'+str(sheet_name)+'Webページから全体データ取得中...')

        print('アクセス成功。'+str(sheet_name)+'Webページから全体データ取得中...')
        html_source = browser.page_source # ブラウザのページソースを取得する
        soup = BeautifulSoup(html_source, 'html.parser') # BeautifulSoupでブラウザのページソースを解析する
        soup_shop = soup.find('ul', class_=lambda x: x and x.startswith('productList')) # ランキングの商品情報が入っているulタグを指定する

        data=[] 
        dig_shops = soup_shop.find_all('div', class_='tileListTtl__txt') # 商品情報が入っているdivタグを指定する

        for index,item in enumerate(dig_shops):
            index_plus = index + 1 #順位が違う場合変える
            print(str(sheet_name)+f'No.{index_plus}のURLを取得中...')

            dig_url_item = item.find('a').get('href')
            dig_url = 'https://www.dmm.co.jp' + dig_url_item

            # Chromeオプションの設定
            chrome_options = Options()
            chrome_options.add_argument("--headless")             # ヘッドレスモードを有効にする
            prefs = {"profile.managed_default_content_settings.images": 2}             # 画像を読み込まない設定を追加
            chrome_options.add_argument('--disable-blinkfeatures=AutomationControlled')
            chrome_options.add_experimental_option("prefs", prefs)
            # ChromeDriverManagerを使用してChromeドライバーを起動する
            browser = webdriver.Chrome(ChromeDriverManager().install(), options=chrome_options)

            browser.get(dig_url) # 指定したURLのページを開く

            try:
                elem_login_btn_top = browser.find_element_by_css_selector('.ageCheck__link--r18') # ログインボタンの要素を指定する
                elem_login_btn_top.click() # ログインボタンをクリックする
            except NoSuchElementException:
                pass

            print('アクセス成功。'+str(sheet_name)+str(index_plus )+'位の個別Webページのデータ取得中')
            html_source = browser.page_source # ブラウザのページソースを取得する
            shop_page_soup = BeautifulSoup(html_source, 'html.parser') # BeautifulSoupでブラウザのページソースを解析する

            print(str(sheet_name)+str(index_plus )+'位の個別Webページのデータ解析中')

            #24時間、週間、月間ランキング情報取得
            ranking_items = shop_page_soup.find_all('span', 'rankingList__txt--number')

            #24時間ランキング
            try:
                ranking_items_day = int(ranking_items[0].text)
            except IndexError:
                ranking_items_day = '―'  # 空欄表記

            #週間ランキング
            try:
                ranking_items_week = int(ranking_items[1].text)
            except IndexError:
                ranking_items_week =  '―'  # 空欄表記

            #月間ランキング
            try:
                ranking_items_month = int(ranking_items[2].text)
            except IndexError:
                ranking_items_month =  '―'  # 空欄表記

            #専売の有無
            FANZA_exclusive = "×"
            try:
                exclusive_items = shop_page_soup.find('span', 'c_icon_exclusive -detail')
                if "DLsite専売" in exclusive_items.text.strip():
                    FANZA_exclusive = "○"
                else:
                    pass
            except Exception as e:
                pass    

            #商品名
            product_title = shop_page_soup.find('h1','productTitle__txt').text.strip()   
            try:
                # re.subで【】とその中の文字を削除。正規表現の【.*?】が【】とその中身にマッチします。
                product_title_str = re.sub(r'【.*?】', '', product_title)
                # 余分な空白や改行を削除
                product_title_str = product_title_str.strip()
            except Exception as e:
                pass

            print('商品名取得。' + str(sheet_name) + str(index_plus ) +'位の商品名は、「' + product_title_str + '」です。')
            
            #新作・旧作表記
            element = shop_page_soup.find('span', 'c_icon_latest')
            if element is not None:
                latest_situation = element.text.strip()
            else:
                latest_situation = '―'

            #ブランド名の取得
            brand_name_element = shop_page_soup.find('a', class_='circleName__txt')
            brand_name_text = brand_name_element.text.strip().replace(",", "")

            #ブランドURL
            brand_name_element_url = brand_name_element.get('href')
            brand_name_element_url
            brand_url = 'https://www.dmm.co.jp'+brand_name_element_url

            #ゲームジャンルの取得
            genre_name_elements = shop_page_soup.find_all('div', class_='productInformation__item')

            for element in genre_name_elements:
                if 'ゲームジャンル' in element.text:
                    genre_name_text = element.find('dd', class_='informationList__txt').text.strip()
                    break
            else:
                genre_name_text  = '―'  # 空欄表記

            #トータル販売数の取得
            number_of_sales = shop_page_soup.find('span', class_='numberOfSales__txt')
            number_of_sales_int = int(number_of_sales.text.strip().replace(",", ""))

            #販売価格
            selling_price = shop_page_soup.find('p', class_=lambda x: x and x.startswith('priceList__main'))
            if  selling_price  is not None:
                selling_price_str = selling_price.text.strip().replace('円', '')
                selling_price_int = int(re.search(r'(\d+)',selling_price_str.replace(',', '')).group(1))
            else:
                selling_price_int  = '―'  # 空欄表記

            #卸売価格ベースの販売予想額
            def calculate_wholesale_price(selling_price_int):
                # 卸売価格の割合を決定する
                if selling_price_int <= 500:
                    ratio = 0.5  # 50%の割合
                elif 500 < selling_price_int <= 1000:
                    ratio = 0.55  # 55%の割合
                elif 1000 < selling_price_int <= 2000:
                    ratio = 0.6  # 60%の割合
                else:
                    ratio = 0.65  # 65%の割合
                # 卸売価格を計算
                wholesale_price = selling_price_int * ratio
                return int(wholesale_price)  # 結果を整数として返す
            wholesale_price_result = calculate_wholesale_price(selling_price_int) * number_of_sales_int #販売予想額

            #割引率
            discount_rate = shop_page_soup.find('p','campaignBalloon__ttl')
            if  discount_rate is not None:
                discount_rate_str = discount_rate.text.strip().replace('円', '')
                discount_rate_int = int(re.search(r'(\d+)',discount_rate_str.replace(',', '')).group(1))
            else:
                discount_rate_int  = '―'  # 空欄表記

            #サークル設定価格
            main_price = shop_page_soup.find('span', class_=lambda x: x and x.startswith('priceList__sub'))
            if  main_price  is not None:
                main_price_str = main_price.text.strip().replace('円', '')
                main_price_int = int(re.search(r'(\d+)',main_price_str.replace(',', '')).group(1))
            else:
                main_price = shop_page_soup.find('p', class_=lambda x: x and x.startswith('priceList__main'))
                main_price_str = main_price.text.strip().replace('円', '')
                main_price_int = int(re.search(r'(\d+)',main_price_str.replace(',', '')).group(1))

            #ジャンルタグの取得
            search_tag = shop_page_soup.find('ul', class_='genreTagList') 
            if search_tag is not None:
                tags = [a_tag.text for a_tag in search_tag.find_all('a')] # aタグのテキスト内容を取得し、一つの文字列に結合
                tags = [tag.strip() for tag in tags] # タグの前後の空白を削除
                output_string = '／'.join(tags) # タグを「／」で区切って一つの文字列に結合
            else:
                output_string = '―'  # 空欄表記
            final_string = output_string.replace("準新作", "").replace("新作", "").replace("旧作", "").replace("成人向け", "").replace("男性向け", "")
            final_string_re = re.sub('／／+', '／', final_string)

            #平均評価
            rewiew_average_item = shop_page_soup.find('p','dcd-review__average')
            if rewiew_average_item is not None:
                rewiew_average = rewiew_average_item.find('strong').text
            else:
                rewiew_average = '―'

            #総評価数
            rewiew_evaluate_item = shop_page_soup.find('p','dcd-review__evaluates')
            if rewiew_evaluate_item is not None:
                rewiew_evaluate = rewiew_evaluate_item.find('strong').text
            else:
                rewiew_evaluate = '―'
        
            #★1～5の数の定義
            rewiew_5star_item = shop_page_soup.find('div','dcd-review__rating_map')
            if rewiew_5star_item is not None:
                rewiew_5star = [int(span.text[:-1]) for span in rewiew_5star_item.find_all('span') if '件' in span.text]
                rewiew_5star_num = rewiew_5star[0]
                rewiew_4star_num = rewiew_5star[1] 
                rewiew_3star_num = rewiew_5star[2] 
                rewiew_2star_num = rewiew_5star[3] 
                rewiew_1star_num = rewiew_5star[4] 
            else:
                rewiew_5star_num = '―'
                rewiew_4star_num = '―'
                rewiew_3star_num = '―'
                rewiew_2star_num = '―'
                rewiew_1star_num = '―'

            #お気に入り登録数
            favorite_evaluate_item = shop_page_soup.find('span','favorites__txt')
            if  favorite_evaluate_item is not None: 
                favorite_number = int(re.sub(r'[^\d,]', '', favorite_evaluate_item.text).replace(',', ''))
            else:
                favorite_number = '―'

            #「ファンになる」登録数
            fan_evaluate_item = shop_page_soup.find('div','circleFanCount__txt')
            if  fan_evaluate_item is not None:
                fan_number = int(re.sub(r'[^\d,]', '', fan_evaluate_item.text).replace(',', ''))
            else:
                fan_number = '―'

            #配信開始日
            release_day_element = None  # 親要素からラベルと値を取得するための変数
            for item in shop_page_soup.find_all('div', class_='productInformation__item'):  # ラベルと値が含まれる親要素を探す
                if '配信開始日' in item.text:
                    release_day_element = item
                    break
            if release_day_element:  # 対応する値を取得する
                date_text = release_day_element.find('dd', class_='informationList__txt').text.strip()
                if re.search(r'\d{4}/\d{2}/\d{2}', date_text):
                    target_date = re.search(r'\d{4}/\d{2}/\d{2}', date_text).group(0)
            else:
                target_date = ''  # 空欄表記

            # 配信後の日数
            now_2 = datetime.now()
            timestamp_2 = now_2.strftime("%Y%m%d")
            date_format_target = "%Y/%m/%d"  # target_dateの形式を変更
            date_format_timestamp = "%Y%m%d"
            target_date_obj = datetime.strptime(target_date, date_format_target)
            timestamp_2_obj = datetime.strptime(timestamp_2, date_format_timestamp)
            pass_day = (timestamp_2_obj - target_date_obj).days

            #ファイル容量
            file_vol_element = None  # 親要素からラベルと値を取得するための変数
            for item in shop_page_soup.find_all('div', class_='productInformation__item'):  # ラベルと値が含まれる親要素を探す
                if 'ファイル容量' in item.text:
                    file_vol_element = item
                    break
            if file_vol_element:  # 対応する値を取得する
                file_vol = file_vol_element.find('dd', class_='informationList__txt').text.strip()
            else:    
                file_vol = ''  # 空欄表記

            # details辞書にデータを格納
            details = {}
            datum=details
            details['No'] = (index_plus)
            details['24\n時間\nﾗﾝ\nｷﾝｸﾞ'] = ranking_items_day
            details['週間\nﾗﾝ\nｷﾝｸﾞ'] = ranking_items_week
            details['月間\nﾗﾝ\nｷﾝｸﾞ'] = ranking_items_month
            details['FANZA\n専売'] = FANZA_exclusive
            details['商品名'] = product_title_str
            details['新旧\n表記'] = latest_situation
            details['ﾌﾞﾗﾝﾄﾞ名'] = brand_name_text 
            details['ゲーム\nｼﾞｬﾝﾙ'] = genre_name_text
            details['ﾄｰﾀﾙ\n販売数'] =  number_of_sales_int
            details['販売\n価格'] = selling_price_int
            details['登録\n売価'] = main_price_int 
            details['割\n引\n率'] = discount_rate_int
            details['予想\n粗利\n額'] = wholesale_price_result 
            details['ｼﾞｬﾝﾙ\nタグ'] = final_string_re
            details['平均\n評価'] = rewiew_average
            details['総評\n価数'] = rewiew_evaluate
            details['★5\nの数'] = rewiew_5star_num
            details['★4\nの数'] = rewiew_4star_num
            details['★3\nの数'] = rewiew_3star_num 
            details['★2\nの数'] = rewiew_2star_num
            details['★1\nの数'] = rewiew_1star_num 
            details['お気に\n入り\n登録数'] = favorite_number 
            details['ファン\nになる\n登録数'] = fan_number 
            details['配信\n開始日'] = target_date
            details['本ﾃﾞｰﾀ\n作成日\nとの差'] = pass_day
            details['ﾌｧｲﾙ\n容量'] =file_vol 
            details['商品URL'] = dig_url
            details['ブランドURL'] = brand_url
            data.append(datum)

            print(str(sheet_name)+str(index_plus )+'位の個別データ書き込み完了……')
            browser.quit()
            sleep(1) 

        df = pd.DataFrame(data)
        df.index = np.arange(1, len(df)+1)
        df.to_excel(writer, sheet_name,index=False)

        # ユーザーのデスクトップパスを取得
        desktop_path = os.path.join(os.path.join(os.environ['USERPROFILE']), 'Desktop')
        # 今日の日付を取得
        today_date = datetime.today().strftime('%Y%m%d')
        # 保存先のフォルダ名を指定（今日の日付を追加）
        folder_name = 'img_' + today_date
        # フルパスを作成
        save_path = os.path.join(desktop_path, folder_name)
        # フォルダが存在しない場合、作成
        if not os.path.exists(save_path):
            os.makedirs(save_path)
            
        # 日本語フォントの設定
        mpl.rcParams['font.family'] = 'Yu Mincho' # 例として「Yu Mincho」を使用
        plt.rcParams['font.family'] = 'Yu Gothic' # もしくは他の日本語対応のフォント名
        
        # 販売数と現状価格のプロット
        plt.figure()
        plt.scatter(df['販売\n価格'],df['ﾄｰﾀﾙ\n販売数'])  # 点をプロット
        plt.title('販売価格と販売数', fontsize=18, fontweight='bold')  # タイトルをゴシック、太字、18ポイントに
        plt.xlabel('販売価格')  # x軸のラベル
        plt.ylabel('販売数')  # y軸のラベル
        plt.savefig(save_path + '\\graph_' + str(sheet_name) + '1.png')
        plt.close()

        plt.figure()
        plt.scatter( df['登録\n売価'],df['販売\n価格'])  # 点をプロット
        plt.title('登録売価と販売価格', fontsize=18, fontweight='bold')
        plt.xlabel('登録売価')  # x軸のラベル
        plt.ylabel('販売価格')  # y軸のラベル
        plt.savefig(save_path + '\\graph_' + str(sheet_name) + '2.png')
        plt.close()

        # サークル数と粗利金額のローレンツ曲線のプロット
        plt.figure()
        plt.scatter(range(len(df['予想\n粗利\n額'])), [x / 10000 for x in sorted(df['予想\n粗利\n額'])])  # 点をプロット、y値を10000で割る
        plt.title('サークル数(累計数)と粗利金額', fontsize=18, fontweight='bold')
        plt.xlabel('サークル数(累計数)')  # x軸のラベル
        plt.ylabel('粗利金額 (万)')    # y軸のラベルに「万」を追加
        plt.ticklabel_format(style='plain', axis='y')  # y軸のラベルを指数表記にしない
        plt.savefig(save_path + '\\graph_' + str(sheet_name) + '3.png')
        plt.close()

        plt.figure()
        plt.scatter(df['本ﾃﾞｰﾀ\n作成日\nとの差'], [x / 10000 for x in sorted(df['予想\n粗利\n額'])])  # 点をプロット
        plt.title('経過日数と粗利金額', fontsize=18, fontweight='bold')
        plt.xlabel('経過日数')  # x軸のラベル
        plt.ylabel('粗利金額 (万)')  # y軸のラベル
        plt.savefig(save_path + '\\graph_' + str(sheet_name) + '4.png')
        plt.close()
        
        browser.quit()

print(f"Excelファイルに書き込み完了……: {filename}")
browser.quit()

print(f"Excelファイルの見栄えを整えます……: {filename}")
#Excelファイル操作
workbook = load_workbook(filename=filename) # Excelファイルを開く
for sheet_name in workbook.sheetnames:
    sheet = workbook[sheet_name]
    # 1行目に空欄行を挿入する
    sheet.insert_rows(0)
    # B列目に空欄列を挿入する
    sheet.insert_cols(1)

    # 2行目の背景を灰色にする
    grey_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    for cell in sheet['2:2']:
        cell.fill = grey_fill

    # 列の幅を広げる (1単位 = 0.1956cm)
    sheet.column_dimensions['A'].width = 0.52 / 0.1956
    sheet.column_dimensions['B'].width = 1.01 / 0.1956
    sheet.column_dimensions['C'].width = 1.01 / 0.1956
    sheet.column_dimensions['D'].width = 1.01 / 0.1956
    sheet.column_dimensions['E'].width = 1.01 / 0.1956
    sheet.column_dimensions['F'].width = 1.01 / 0.1956
    sheet.column_dimensions['G'].width = 6.40 / 0.1956
    sheet.column_dimensions['H'].width = 1.41 / 0.1956
    sheet.column_dimensions['I'].width = 5.6 / 0.1956
    sheet.column_dimensions['J'].width = 3.68 / 0.1956
    sheet.column_dimensions['K'].width = 1.39 / 0.1956
    sheet.column_dimensions['L'].width = 1.39 / 0.1956  
    sheet.column_dimensions['M'].width = 1.39 / 0.1956
    sheet.column_dimensions['N'].width = 1.12 / 0.1956
    sheet.column_dimensions['O'].width = 2.23 / 0.1956
    sheet.column_dimensions['P'].width = 8.20 / 0.1956
    sheet.column_dimensions['Q'].width = 1.06 / 0.1956
    sheet.column_dimensions['R'].width = 1.06 / 0.1956
    sheet.column_dimensions['S'].width = 0.92 / 0.1956
    sheet.column_dimensions['T'].width = 0.92 / 0.1956
    sheet.column_dimensions['U'].width = 0.92 / 0.1956
    sheet.column_dimensions['V'].width = 0.92 / 0.1956
    sheet.column_dimensions['W'].width = 0.92 / 0.1956
    sheet.column_dimensions['X'].width = 1.61 / 0.1956
    sheet.column_dimensions['Y'].width = 1.61 / 0.1956
    sheet.column_dimensions['Z'].width = 2.48 / 0.1956
    sheet.column_dimensions['AA'].width = 1.47 / 0.1956
    sheet.column_dimensions['AB'].width = 1.91 / 0.1956
    sheet.column_dimensions['AC'].width = 10.76 / 0.1956
    sheet.column_dimensions['AD'].width = 11.82 / 0.1956

    # 行の幅を広げる
    sheet.row_dimensions[2].height = 12 / 0.1956 

    # 3～102行目の全てのセルに対して、中央揃えを適用
    for row in sheet.iter_rows(min_row=3, max_row=102):
        for cell in row:
            cell.alignment = Alignment(horizontal='center') # 中央揃えのスタイルを適用

    #左揃えに設定
    for col in ['F','G','H','I','O','AB']:
        for cell in sheet[col]:
            cell.alignment = Alignment(horizontal='left') # 左揃えのスタイルを適用
    
    #右揃えに設定
    for col in ['J','K','L','M','N','P','Q','W','X','Z','AA']:
        for cell in sheet[col]:
            cell.alignment = Alignment(horizontal='right') # 右揃えのスタイルを適用

    # 2行目の各セルに対して、縦・横方向の中央揃え
    for cell in sheet['2:2']:
        cell.alignment = Alignment(vertical='center')
        cell.alignment = Alignment(horizontal='center') 

    # 細い罫線のスタイルを定義
    thin_border = Border(left=Side(style='thin'), 
                        right=Side(style='thin'), 
                        top=Side(style='thin'), 
                        bottom=Side(style='thin'))

    # B3からO102までのセルに細い罫線を引く
    for row in sheet.iter_rows(min_row=3, min_col=2, max_row=122, max_col=30):
        for cell in row:
            cell.border = thin_border

    # ハイパーリンク用のスタイルを定義
    for idx, url in enumerate(df['商品URL'], start=3):  # インデックス3から開始 (1,2行目はヘッダー)
        sheet.cell(row=idx, column=29, value=url).hyperlink = url  # AK列に対応する列インデックスは37

    for idx, url in enumerate(df['ブランドURL'], start=3):  # インデックス3から開始 (1,2行目はヘッダー)
        sheet.cell(row=idx, column=30, value=url).hyperlink = url  # AK列に対応する列インデックスは37

    # 画像を挿入
    img1 = px.drawing.image.Image(save_path +'\graph_' + sheet_name + '1.png')
    img1.anchor = 'B124'
    sheet.add_image(img1)

    img2 = px.drawing.image.Image(save_path +'\graph_' + sheet_name + '2.png')
    img2.anchor = 'I124' # 位置を調整する必要があります
    sheet.add_image(img2)

    img3 = px.drawing.image.Image(save_path +'\graph_' + sheet_name + '3.png')
    img3.anchor = 'B149' # 位置を調整する必要があります
    sheet.add_image(img3)

    img4 = px.drawing.image.Image(save_path +'\graph_' + sheet_name + '4.png')
    img4.anchor = 'I149' # 位置を調整する必要があります
    sheet.add_image(img4)

# 変更を保存する
workbook.save(filename)

print(f"Excelファイルの見栄えを整えました。全工程完了: {filename}")